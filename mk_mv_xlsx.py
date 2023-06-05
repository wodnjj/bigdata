from openpyxl import Workbook
import func_movie

# 엑셀파일 쓰기
write_wb = Workbook()

# 이름이 있는 시트를 생성
# write_ws = write_wb.create_sheet('생성시트')

# Sheet1에다 입력
write_ws = write_wb.active
write_ws['A1'] = '순위'
write_ws['B1'] = '제목'
write_ws['C1'] = '평점'
write_ws['D1'] = '예매율'
write_ws['E1'] = '개봉날짜'

m_list = func_movie.get_dmv()
for m in m_list:
    #행 단위로 추가
    write_ws.append(m)

#셀 단위로 추가
# write_ws.cell(5, 5, '5행5열')
write_wb.save('함수로가져온영화순위.xlsx')